#!/usr/bin/env python
# -*- coding: utf-8 -*-
import mysql.connector
from openpyxl import load_workbook
import codecs
import json

mydb = mysql.connector.connect(
  host="localhost",
  user="amir",
  passwd="1234",
  # charset='utf8'
)
print(mydb)
mycursor = mydb.cursor()
mycursor.execute("CREATE DATABASE amir_mahdi_db_project")
mydb = mysql.connector.connect(
  host="localhost",
  user="amir",
  passwd="1234",
  charset='utf8',
  database="amir_mahdi_db_project"
)
mycursor = mydb.cursor()

##create tables

#accesories
# mycursor.execute("create table book (id int primary key, product_title varchar(255), brand_name_en varchar(255), topic varchar(255), publisher varchar(255), pages int)")

#book
mycursor.execute("create table books (id int primary key, product_title varchar(255), brand_name_en varchar(255), topic varchar(255), publisher varchar(255), pages int)")
#puzzle
# mycursor.execute("create table book (id int primary key, product_title varchar(255), brand_name_en varchar(255), topic varchar(255), publisher varchar(255), pages int)")
#mouse
# mycursor.execute("create table book (id int primary key, product_title varchar(255), brand_name_en varchar(255), topic varchar(255), publisher varchar(255), pages int)")

#keyboard
# mycursor.execute("create table book (id int primary key, product_title varchar(255), brand_name_en varchar(255), topic varchar(255), publisher varchar(255), pages int)")


##read data
# raw_target_categories  = ['کتاب چاپی','کیبورد (صفحه کلید)','ماوس (موشواره)','پازل','کیف و کاور گوشی','محافظ صفحه نمایش گوشی']
raw_target_categories  = ['کتاب چاپی']
target_categories = []
category_attrs = {}
for category in raw_target_categories:
    target_categories.append(category.decode('utf-8'))
    category_attrs[category.decode('utf-8')] = []
wb = load_workbook(filename = 'data/product-list.xlsx')
ws = wb.active
num_of_rows = 10001
for i in range(2,num_of_rows):
    #reading i-th row of data
    print(i)
    row_topic = ''
    row_publisher = ''
    row_pages = 0
    row_category = ws['F'+str(i)].value
    row_id = ws['A'+str(i)].value
    row_brand_en = ws['I'+str(i)].value
    row_product_title = ws['B'+str(i)].value
    if  row_category in target_categories:
        try:
            object = json.loads(ws['J'+str(i)].value)
            for key_value in object:
                ##
                if key_value['Key'] == 'موضوع'.decode('utf-8'):
                    row_topic = key_value['Value']
                if key_value['Key'] == 'ناشر'.decode('utf-8'):
                    row_publisher = key_value['Value']
                if key_value['Key'] == 'تعداد صفحات'.decode('utf-8'):
                    row_page = int(key_value['Value'])
            sql = "insert into books(id,product_title,brand_name_en,topic,publisher,pages) value(%i,%s,%s,%s,%s,%i)"
            vale = (row_id,row_product_title,row_brand_en,row_topic,row_publisher,row_pages)
            mycursor.execute(sql, val)
            mydb.commit()
        except ValueError:
            print('Error during parsing json!')
