#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import codecs
import json

raw_target_categories  = ['کتاب چاپی','کیبورد (صفحه کلید)','ماوس (موشواره)','پازل','کیف و کاور گوشی','محافظ صفحه نمایش گوشی']
target_categories = []
# print raw_target_categories
for category in raw_target_categories:
    # print category.decode('utf-8')
    target_categories.append(category.decode('utf-8'))
# print target_categories
coresspondingAttrs = {}
countNumOfAttrs = {}
for category in target_categories:
    coresspondingAttrs[category] = set()

wb = load_workbook(filename = 'data/product-list.xlsx')
ws = wb.active
columns_str = 'ABCDEFGHIJ'
num_of_rows = 10001
for ch in columns_str:
    print(ws[ch+'1'].value)
fileWriter = codecs.open("attributes.txt", "w",'utf-8')
# extract all keys from attr colmumn of target_categories
for i in range(2,num_of_rows):
    # fileWriter.write(ws['F'+str(i)].value+'\n\r')
    print(i)
    if ws['F'+str(i)].value in target_categories:
        # print(object)
        try:
            object = json.loads(ws['J'+str(i)].value)
            # fileWriter.write(str(object))
            for key in object:
                # print(coresspondingAttrs[ws['F'+str(i)].value])
                coresspondingAttrs[ws['F'+str(i)].value].add(key['Key'])
                if ws['F'+str(i)].value+key['Key'] in countNumOfAttrs:
                    countNumOfAttrs[ws['F'+str(i)].value+key['Key']] = countNumOfAttrs[ws['F'+str(i)].value+key['Key']] + 1
                else:
                    countNumOfAttrs[ws['F'+str(i)].value+key['Key']] = 1
        except ValueError:
            print('Error during parsing json!')
for category in target_categories:
    # print('\r\n'+ category + ':\r\n')
    fileWriter.write('\n\r' + category + ':\n\r')
    # print(coresspondingAttrs[category])
    for item in coresspondingAttrs[category]:
        # print(item)
        fileWriter.write(item + ': ' + str(countNumOfAttrs[category+item]) +' ,')
    # print('\r\n-------\r\n')
    fileWriter.write('\r\n------\r\n')
fileWriter.close()
